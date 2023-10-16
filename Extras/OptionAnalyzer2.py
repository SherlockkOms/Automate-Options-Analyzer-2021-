import numpy as np
import csv
import pandas as pd
import datetime
from datetime import date
from nsepy import get_history as gh
from nsepython import *
import json
import os
from openpyxl import load_workbook,cell
import sys
# from datetime import datetime

import openpyxl #Connect the library
from openpyxl.styles.borders import Border, Side, BORDER_THICK #borders
from openpyxl import Workbook
from openpyxl.styles import numbers
from openpyxl.styles import PatternFill#Connect cell styles
from openpyxl.workbook import Workbook
from openpyxl.styles import Font,Alignment,Fill #Connect styles for text
from openpyxl.styles import colors
#added 
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension,DimensionHolder
from openpyxl.utils import get_column_letter
from dateutil.relativedelta import relativedelta, TH

f = open("sharelist.txt", "r")
lines = (line.rstrip() for line in f) # All lines including the blank ones
Lines = (line for line in lines if line) # Non-blank lines
count = 0
# b = np.zeros([20, 25], dtype=object)
# Strips the newline character
td = datetime.datetime.now().strftime("%d-%m-%Y-%H-%M-%S")
fname = "errors_"+td+".log"
logf = open(fname, "w")


b = []
stock_names = []
def build_matrix(cols):
    b.append([0 for c in range(0, cols)])

    return b
for line in Lines:
    #print("Line{}: {}".format(count, line.strip()))
    a = line.strip()
    print(a.split(','))
    build_matrix(24)
    b[count][0] = a.split(",")[0]
    stock_names.append(a.split(",")[1].strip().upper())
    count += 1
f.close()
num_shares = count


f = open("config.txt", "r")
Lines = f.readlines()
# Strips the newline character
count = 0
label = []
value = []

    
for line in Lines:
    #print("Line{}: {}".format(count, line.strip()))
    if line.startswith("#"):
        continue
    else:
        a = line.strip()
        label.append(a.split("=")[0].strip())
        value.append(a.split("=")[1].strip())


curr_month = value[0]
how_far = value[1]
margin = value[2]
buffer_days = int(value[3])

if int(curr_month)>1:
    sys.exit("Current Month out of range!!")

if int(how_far)>100 or int(how_far)<0 :
    sys.exit("How far out of range!!")

if int(margin)>100 or int(margin)<0 :
    sys.exit("Margin out of range!!")


end_of_month = date.today() + relativedelta(day=31,months=int(curr_month))
last_thursday = end_of_month + relativedelta(weekday=TH(-1))

today = date.today()



if last_thursday.day - today.day>=0:
    days_left = (last_thursday - date.today()).days
    if days_left ==0:
        days_left = 1
else:
    last_thursday = end_of_month + relativedelta(weekday=TH(-1),months=1)
    days_left = (last_thursday - date.today()).days


days_left = days_left + buffer_days  
    
for i in range(0,3):
    print(label[i],":",value[i])
print("days_left:",days_left)
print(end_of_month)
print(last_thursday)

print(num_shares)

df_new=pd.read_excel('Holidays.xlsx')
holiday_list = df_new['Date'].astype(str)
holiday_list=holiday_list.tolist()



# today = datetime.date(2021, 7, 4)
print("Today's date:", today)
week_num = today.weekday()
print(week_num)
# week num is 0 to 6.. if week num is 0,then it is a monday, then previous days are off days
todayStr1 = today.strftime("%Y-%m-%d")
while week_num > 4 or todayStr1 in holiday_list:
    print("I am in Today while")
    if week_num > 4:
        print(" It is a weekend Chill")
        today = today-datetime.timedelta(days=1)
        todayStr1 = today.strftime("%Y-%m-%d")
        week_num = week_num -1
    if todayStr1 in holiday_list:
        today = today-datetime.timedelta(days=1)
        todayStr1 = today.strftime("%Y-%m-%d")
        week_num=week_num-1
todayStr = today.strftime("%d,%m,%Y")
print(todayStr)
cd = todayStr.split(",")
print(cd)
ywn = week_num - 1 
day = 1
print(ywn)

if ywn < 0 :
    print("it  is a weekend")
    ywn = 4
    day = 3

yesterday = today - datetime.timedelta(days=day)
print('Yesterday is : ',yesterday)
yesterdayStr1 = yesterday.strftime("%Y-%m-%d")
while yesterdayStr1 in holiday_list:
    print("I am in Yesterday while")
    if yesterdayStr1 in holiday_list:
        yesterday = yesterday-datetime.timedelta(days=1)
        yesterdayStr1 = yesterday.strftime("%Y-%m-%d")
        ywn=ywn - 1
    if ywn < 0 :
        print("it  is a weekend")
        ywn = 4
        yesterday = yesterday-datetime.timedelta(days=2)
        yesterdayStr1 = yesterday.strftime("%Y-%m-%d")

yesterdayStr = yesterday.strftime("%d,%m,%Y")
print(yesterdayStr)
yd = yesterdayStr.split(",")
print(yd)
dywn = ywn - 1
day = 1
print(dywn)
if dywn < 0 :
    print("it  is a weekend")
    dywn = 4
    day = 3


daybefyesterday = yesterday - datetime.timedelta(days=day)
print(daybefyesterday)
daybefyesterdayStr1 = daybefyesterday.strftime("%Y-%m-%d")
while daybefyesterdayStr1 in holiday_list:
    print("I am in Daybefyesterday while")
    if daybefyesterdayStr1 in holiday_list:
        daybefyesterday = daybefyesterday-datetime.timedelta(days=1)
        daybefyesterdayStr1 = daybefyesterday.strftime("%Y-%m-%d")
        dywn=dywn - 1
    if dywn < 0 :
        print("it  is a weekend")
        dywn = 4
        daybefyesterday = daybefyesterday-datetime.timedelta(days=2)
        daybefyesterdayStr1 = daybefyesterday.strftime("%Y-%m-%d")
        
daybefyesterdayStr = daybefyesterday.strftime("%d,%m,%Y")
print(daybefyesterdayStr)
dyd = daybefyesterdayStr.split(",")
print(dyd)
# quit()
last90days = yesterday - datetime.timedelta(days=90)
last90daystr = last90days.strftime("%d,%m,%Y")
print(last90daystr)
l90d = last90daystr.split(",")
print(l90d)

cols_color1 = np.zeros(num_shares)
cols_color2 = np.zeros(num_shares)
diff11 = np.zeros(num_shares*3).reshape(num_shares,3)
diff22 = np.zeros(num_shares*3).reshape(num_shares,3)  
bid1 =  np.zeros(num_shares*3).reshape(num_shares,3)
bid2 = np.zeros(num_shares*3).reshape(num_shares,3)
ask1 = np.zeros(num_shares*3).reshape(num_shares,3)
ask2 = np.zeros(num_shares*3).reshape(num_shares,3)
sp1 = np.zeros(num_shares*3).reshape(num_shares,3)
sp2 = np.zeros(num_shares*3).reshape(num_shares,3)





for i in range(num_shares):
    print("num of share", num_shares)

    print ("Getting for share ==>", b[i][0])
    
    stk1 = gh(symbol=stock_names[i], start=date(int(dyd[2]), int(dyd[1]), int(
    dyd[0])), end=date(int(yd[2]), int(yd[1]), int(yd[0])))

    if len(stk1)==0:
        logf.write("No data found for : {0}".format( str(stock_names[i])))
        #print("data not found for ",stock_names[i]) 
        continue
    
        
    #print(stk1)
    tdf = pd.DataFrame(stk1)
    print("#####")
    try:
        print(tdf.Close[0])
        print(tdf.Close[1])
    except:
        logf.write("\nThere is some error getting close value for {0} ".format(str(b[i][0])))
        #print ("There is some error getting close value for ",b[i][0])
        continue

    b[i][2] = tdf.Close[1]
    b[i][3] = tdf.Close[0]

    print(b)

    print("####")
    print("Getting the nearest option value")
    s = nse_optionchain_scrapper(stock_names[i])
    if len(s) == 0:
        logf.write("\nNo data found for : {0}".format( str(b[i][0])))
        #print("No data found for :", b[i][0])
        continue
        
    #type(s)
    #print (s)
    



    #quit ()
    ltp = s["records"]["underlyingValue"]
    timestamp = s["records"]["timestamp"]
    print("Last Traded Price = ", ltp)
    b[i][1] = ltp
    print("Time stamp at which NSE scrapper was run", timestamp)
    last_sp = (s["records"]["data"][1]["strikePrice"])
    last_to_lastsp = (s["records"]["data"][0]["strikePrice"])

    print("#####")
    b[i][4] = ltp   # Nearest Option
    diff00 = np.zeros(num_shares*len(s["records"]["data"])).reshape(num_shares,len(s["records"]["data"]))
    #print(diff00.shape)
    #print(i)
    

    for x in range(0,len(s["records"]["data"])):
        if last_thursday.strftime("%d-%b-%Y") != s["records"]["data"][x]["expiryDate"]:
            diff00[i][x] = 9999

        if last_thursday.strftime("%d-%b-%Y") == s["records"]["data"][x]["expiryDate"]:
            diff00[i][x] = np.absolute(b[i][4]-s["records"]["data"][x]["strikePrice"])
    arg = np.argmin(diff00[i])
    b[i][4] = s["records"]["data"][arg]["strikePrice"]
    try:

        b[i][5] = round(s["records"]["data"][arg]["CE"]["bidprice"],2)
    except:
        logf.write("\nNo CE for : {0} share {1}".format( str(b[i][4]),b[i][0]))
        #print("no CE")
    try:
        b[i][6] = round(s["records"]["data"][arg]["PE"]["bidprice"],2)
    except:
        logf.write("\nNo PE for : {0} and share {1}".format( str(b[i][4]),b[i][0]))
        #print("no PE")


    b[i][7] = round((b[i][5] + b[i][6]), 2)    #PREMIUM TOTAL
    b[i][8] = round(((b[i][7] / b[i][2]) * 100), 2)      # AB IRR
    b[i][9] = round((((float(b[i][7]) / float(b[i][4]) /int(days_left) * 365 / (int(margin)) / 2))*100)*100, 2)    # IRR PA
    b[i][10] = b[i][4] - b[i][7]
    b[i][11] = b[i][4] + b[i][7]

    b[i][12] = round(b[i][7] / (b[i][4] - b[i][7])*100, 2)
    b[i][13] = round(b[i][7] / (b[i][4] + b[i][7])*100, 2)
    
    b[i][16] = float(b[i][4] * (1 - int(how_far)/100))
    b[i][20] = float(b[i][4] * (1 + int(how_far)/100))
    print("12% higher lower values are")
    print(b[i][16],b[i][20])
    diff16 = np.zeros(num_shares*len(s["records"]["data"])).reshape(num_shares,len(s["records"]["data"]))
    diff20 = np.zeros(num_shares*len(s["records"]["data"])).reshape(num_shares,len(s["records"]["data"]))
    
    for x in range(0, len(s["records"]["data"])):
        #print(s["records"]["data"][x]["strikePrice"])
        if last_thursday.strftime("%d-%b-%Y") != s["records"]["data"][x]["expiryDate"]:
            diff16[i][x] = 9999
            diff20[i][x] = 9999

        if last_thursday.strftime("%d-%b-%Y") == s["records"]["data"][x]["expiryDate"]:
            diff16[i][x] = np.absolute(b[i][16]-s["records"]["data"][x]["strikePrice"])
            diff20[i][x] = np.absolute(b[i][20]-s["records"]["data"][x]["strikePrice"])
    #print("DIFF16",diff16)
    #print("DIFF20",diff20)
    idx1 = np.zeros(1)
    idx2 = np.zeros(1)
    idx1 = np.argsort(diff16[i])[0]
    idx2 = np.argsort(diff20[i])[0]
    print ("IDX",idx1)
    print("IDX2",idx2)
    b[i][16] = s["records"]["data"][int(idx1)]["strikePrice"]
    b[i][20] = s["records"]["data"][int(idx2)]["strikePrice"]
    
    print("post adjusetment values",b[i][16],b[i][20])

     
    print("#####")
    print("#####")


    print(stock_names[i])
    if len(s) == 0:
        logf.write("\nNo data found for : {0}".format( str(b[i][0])))
        #print("No data found for :", b[i][0])
        continue
    else:
        if float(b[i][16]) <= float(s["records"]["data"][0]["strikePrice"]):
                    b[i][16] = s["records"]["data"][0]["strikePrice"]
        else:
            b[i][16] = b[i][16]

        if b[i][16] == s["records"]["data"][0]["strikePrice"]:
            for j in range(0,3):
                if b[i][16] == s["records"]["data"][j]["strikePrice"] and (last_thursday.strftime("%d-%b-%Y") == s["records"]["data"][j]["expiryDate"]):
                    try:
                        b[i][17] = s["records"]["data"][j]["PE"]["bidprice"]
                        turns=1
                        cols_color1[i] = turns
                    except:
                        logf.write("No PE Data found for {0}".format(b[i][0]))
                    try:
                        ratio2 = round(s["records"]["data"][j]["PE"]["askPrice"]- s["records"]["data"][j]["PE"]["bidprice"],2)
                    except:
                        logf.write("\nNo PE for : {0} and share {1}".format( str(s["records"]["data"][0]['strikePrice']),b[i][0]))
                        #print("No PE found")
                        ratio2 = 0 

                diff11[i][turns-1] = ratio2
                print(b[i][17],cols_color1[i],diff11[i][turns-1])
        else:
            count = 1
            for x in (range(idx1,idx1+3)):
                if int(b[i][16]) == int(s["records"]["data"][x]["strikePrice"]) and(last_thursday.strftime("%d-%b-%Y") == s["records"]["data"][x]["expiryDate"]):
                    try:
                        last_price = s["records"]["data"][x]["PE"]["bidprice"]
                    except:
                        last_price = 9999
                        logf.write("\nNo PE for : {0} and shhare {1}".format( str(s["records"]["data"][x]['strikePrice']),b[i][0]))
                        continue
                        #print("NO PE DATA")
                    if s["records"]["data"][x]["PE"]["askPrice"]==0:
                        ratio1 = 9999
                    else:
                        ratio1 = (s["records"]["data"][x]["PE"]["askPrice"]- s["records"]["data"][x]["PE"]["bidprice"])/s["records"]["data"][x]["PE"]["askPrice"]
                    ratio2 = round(s["records"]["data"][x]["PE"]["askPrice"]- s["records"]["data"][x]["PE"]["bidprice"],2)
                    print("BID PRICE , RATIO1, RATIO2")
                    print(last_price,ratio1,ratio2)

                    if ratio1<=0.15 :
                        b[i][17]= last_price
                        cols_color1[i] = count
                        diff11[i][count-1] = ratio2
                        bid1[i][count-1] = s["records"]["data"][x]["PE"]["bidprice"]
                        ask1[i][count-1] = s["records"]["data"][x]["PE"]["askPrice"]
                        sp1[i][count-1] = s["records"]["data"][x]["PE"]["strikePrice"]
                        break
                            
                            
                    else:
                        diff11[i][count-1] = ratio2
                        bid1[i][count-1] = s["records"]["data"][x]["PE"]["bidprice"]
                        ask1[i][count-1] = s["records"]["data"][x]["PE"]["askPrice"]    
                        sp1[i][count-1] = s["records"]["data"][x]["PE"]["strikePrice"]
                        print("Cant be done in turn",count)
                        print("OLD B16",b[i][16])
                        count = count + 1
                        break
            if count==2:
                for z in range(1,(len(s["records"]["data"])-idx1)):
                    if idx1+z > len(s["records"]["data"]):
                        logf.write("PE Out of bounds for {0}".format(b[i][0]))
                        break
                    else:           
                        if s["records"]["data"][idx1+z]["expiryDate"] == last_thursday.strftime("%d-%b-%Y"):
                            print("changing b16")
                            b[i][16] = s["records"]["data"][idx1 + z]["strikePrice"]
                            print("NEW B16",b[i][16])
                            try:
                                last_price = s["records"]["data"][idx1 + z]["PE"]["bidprice"]
                            except:
                                last_price = 9999
                                logf.write("\nNo PE for : {0} and share".format( str(s["records"]["data"][idx1 + z]['strikePrice']),b[i][0]))
                                continue
                                #print("NO PE DATA")
                            if s["records"]["data"][idx1 + z]["PE"]["askPrice"]==0:
                                ratio1 = 9999
                            else:
                                ratio1 = (s["records"]["data"][idx1 + z]["PE"]["askPrice"]- s["records"]["data"][idx1 + z]["PE"]["bidprice"])/s["records"]["data"][idx1 + z]["PE"]["askPrice"]
                            ratio2 = round(s["records"]["data"][idx1+z]["PE"]["askPrice"]- s["records"]["data"][idx1+z]["PE"]["bidprice"],2)
                            print("BID PRICE , RATIO1, RATIO2")
                            print(last_price,ratio1,ratio2)

                            if ratio1<=0.15 :
                                b[i][17]= last_price
                                cols_color1[i] = count
                                diff11[i][count-1] = ratio2
                                bid1[i][count-1] = s["records"]["data"][idx1 + z]["PE"]["bidprice"]
                                ask1[i][count-1] = s["records"]["data"][idx1 + z]["PE"]["askPrice"]
                                sp1[i][count-1] = s["records"]["data"][idx1+z]["PE"]["strikePrice"]
                                break
                                            
                            else:
                                diff11[i][count-1] = ratio2
                                bid1[i][count-1] = s["records"]["data"][idx1 + z]["PE"]["bidprice"]
                                ask1[i][count-1] = s["records"]["data"][idx1 + z]["PE"]["askPrice"]  
                                sp1[i][count-1] = s["records"]["data"][idx1+z]["PE"]["strikePrice"]  
                                print("Cant be done in turn",count)
                                print("OLD B16",b[i][16])
                                count = count+1
                                break
                    

            if count==3:
                for z in range(1,idx1):
                    if idx1-z < 0:
                        logf.write("PE Out of bounds for {0}".format(b[i][0]))
                        break
                    else:
                        if s["records"]["data"][idx1-z]["expiryDate"] == last_thursday.strftime("%d-%b-%Y"):
                            print("changing b16")
                            b[i][16] = s["records"]["data"][idx1 - z]["strikePrice"]
                            print("NEW B16",b[i][16])
                            try:
                                last_price = s["records"]["data"][idx1 - z]["PE"]["bidprice"]
                            except:
                                last_price = 9999
                                logf.write("\nNo PE for : {0} and share{1}".format( str(s["records"]["data"][idx1 - z]['strikePrice']),b[i][0]))
                                continue
                                #print("NO PE DATA")
                            if s["records"]["data"][idx1 - z]["PE"]["askPrice"]==0:
                                ratio1 = 9999
                            else:
                                ratio1 = (s["records"]["data"][idx1 - z]["PE"]["askPrice"]- s["records"]["data"][idx1 - z]["PE"]["bidprice"])/s["records"]["data"][idx1 - z]["PE"]["askPrice"]
                            ratio2 = round(s["records"]["data"][idx1-z]["PE"]["askPrice"]- s["records"]["data"][idx1-z]["PE"]["bidprice"],2)
                            print("BID PRICE , RATIO1, RATIO2")
                            print(last_price,ratio1,ratio2)

                            if ratio1<=0.15 :
                                b[i][17]= last_price
                                cols_color1[i] = count
                                diff11[i][count-1] = ratio2
                                bid1[i][count-1] = s["records"]["data"][idx1-z]["PE"]["bidprice"]
                                ask1[i][count-1] = s["records"]["data"][idx1-z]["PE"]["askPrice"]
                                sp1[i][count-1] = s["records"]["data"][idx1-z]["PE"]["strikePrice"]
                                break
                                            
                            else:
                                diff11[i][count-1] = ratio2
                                bid1[i][count-1] = s["records"]["data"][idx1-z]["PE"]["bidprice"]
                                ask1[i][count-1] = s["records"]["data"][idx1-z]["PE"]["askPrice"]  
                                sp1[i][count-1] = s["records"]["data"][idx1-z]["PE"]["strikePrice"]  
                                print("Cant be done in turn",count)
                                print("OLD B16",b[i][16])
                                #print(diff11)
                                cols_color1[i] = count+1
                                result = np.all(diff11[i] == diff11[i][0])
                                if result:
                                    b[i][16] = s["records"]["data"][int(idx1)]["strikePrice"]
                                    m2 = np.mean([bid1[i][0],ask1[i][0]])
                                    b[i][17] = float('%.3f'%m2)
                                else:
                                    diff11[i][diff11[i]<=0] = 9999 
                                    b[i][16] = sp1[i][np.argmin(diff11[i])]
                                    m3 = np.mean([ask1[i][np.argmin(diff11[i])], bid1[i][np.argmin(diff11[i])]])
                                    b[i][17] = float('%.3f'%m3)
                                    print("avg is",b[i][17])
                                    print("DIFF IS " ,diff11[i])
                                    print("strike price selected is",sp1[i][np.argmin(diff11[i])])
                                    print("ask and bid are",ask1[i][np.argmin(diff11[i])],bid1[i][np.argmin(diff11[i])])
                                break
                    
                    
        print("BI20 =",b[i][20])                
        if float(b[i][20]) > float(s["records"]["data"][len(s["records"]["data"])-1]["strikePrice"]):
                    b[i][20] = s["records"]["data"][len(s["records"]["data"])-1]["strikePrice"]
        else:
            b[i][20] = b[i][20]
        
        
        if b[i][20] == s["records"]["data"][len(s["records"]["data"])-1]["strikePrice"]:
            for j in range(len(s["records"]["data"])-3,len(s["records"]["data"])):
                if b[i][20] == s["records"]["data"][j]["strikePrice"] and (last_thursday.strftime("%d-%b-%Y") == s["records"]["data"][j]["expiryDate"]):
                    try:
                        b[i][21] = s["records"]["data"][j]["CE"]["bidprice"]
                        turns=1
                        cols_color2[i] = turns
                    except:
                        logf.write("No CE Data found for {0}".format(b[i][0]))
        else:
            count = 1
            for x in (range(idx2,idx2+3)):
                if int(b[i][20]) == int(s["records"]["data"][x]["strikePrice"]) and(last_thursday.strftime("%d-%b-%Y") == s["records"]["data"][x]["expiryDate"]):
                    try:
                        last_price = s["records"]["data"][x]["CE"]["bidprice"]
                    except:
                        last_price = 9999
                        logf.write("\nNo CE for : {0} share {1}".format( str(s["records"]["data"][x]['strikePrice']),b[i][0]))
                        continue
                        #print("NO PE DATA")
                    if s["records"]["data"][x]["CE"]["askPrice"]==0:
                        ratio1 = 9999
                    else:
                        ratio1 = (s["records"]["data"][x]["CE"]["askPrice"]- s["records"]["data"][x]["CE"]["bidprice"])/s["records"]["data"][x]["CE"]["askPrice"]
                    ratio2 = round(s["records"]["data"][x]["CE"]["askPrice"]- s["records"]["data"][x]["CE"]["bidprice"],2)
                    print("BID PRICE , RATIO1, RATIO2")
                    print(last_price,ratio1,ratio2)

                    if ratio1<=0.15 :
                        b[i][21]= last_price
                        cols_color2[i] = count
                        diff22[i][count-1] = ratio2
                        bid2[i][count-1] = s["records"]["data"][x]["CE"]["bidprice"]
                        ask2[i][count-1] = s["records"]["data"][x]["CE"]["askPrice"]
                        sp2[i][count-1] = s["records"]["data"][x]["CE"]["strikePrice"]
                        break
                            
                            
                    else:
                        diff22[i][count-1] = ratio2
                        bid2[i][count-1] = s["records"]["data"][x]["CE"]["bidprice"]
                        ask2[i][count-1] = s["records"]["data"][x]["CE"]["askPrice"]    
                        sp2[i][count-1] = s["records"]["data"][x]["CE"]["strikePrice"]
                        print("Cant be done in turn",count)
                        print("OLD B20",b[i][20])
                        count = count + 1
                        break
            if count==2:
                for z in range(1,idx2):
                    if idx2-z < 0:
                        logf.write("\nCE Out of bounds for {0}".format(b[i][0]))
                        break
                    else:           
                        if s["records"]["data"][idx2-z]["expiryDate"] == last_thursday.strftime("%d-%b-%Y"):
                            print("changing b20")
                            b[i][20] = s["records"]["data"][idx2-z]["strikePrice"]
                            print("NEW B20",b[i][20])
                            try:
                                last_price = s["records"]["data"][idx2-z]["CE"]["bidprice"]
                            except:
                                last_price = 9999
                                logf.write("\nNo CE for : {0} share {1}".format( str(s["records"]["data"][idx2-z]['strikePrice']),b[i][0]))
                                continue
                                #print("NO PE DATA")
                            if s["records"]["data"][idx2-z]["CE"]["askPrice"]==0:
                                ratio1 = 9999
                            else:
                                ratio1 = (s["records"]["data"][idx2-z]["CE"]["askPrice"]- s["records"]["data"][idx2-z]["CE"]["bidprice"])/s["records"]["data"][idx2-z]["CE"]["askPrice"]
                            ratio2 = round(s["records"]["data"][idx2-z]["CE"]["askPrice"]- s["records"]["data"][idx2-z]["CE"]["bidprice"],2)
                            print("BID PRICE , RATIO1, RATIO2")
                            print(last_price,ratio1,ratio2)

                            if ratio1<=0.15 :
                                b[i][21]= last_price
                                cols_color2[i] = count
                                diff22[i][count-1] = ratio2
                                bid2[i][count-1] = s["records"]["data"][idx2-z]["CE"]["bidprice"]
                                ask2[i][count-1] = s["records"]["data"][idx2-z]["CE"]["askPrice"]
                                sp2[i][count-1] = s["records"]["data"][idx2-z]["CE"]["strikePrice"]
                                break
                                            
                            else:
                                diff22[i][count-1] = ratio2
                                bid2[i][count-1] = s["records"]["data"][idx2-z]["CE"]["bidprice"]
                                ask2[i][count-1] = s["records"]["data"][idx2-z]["CE"]["askPrice"]   
                                sp2[i][count-1] = s["records"]["data"][idx2-z]["CE"]["strikePrice"] 
                                print("Cant be done in turn",count)
                                print("OLD B20",b[i][20])
                                count = count+1
                                break
                    

            if count==3:
                for z in range(1,(len(s["records"]["data"])-idx2)):
                    if idx2+z > len(s["records"]["data"]):
                        logf.write("\nCE Out of bounds for {0}".format(b[i][0]))
                        break
                    else:
                    
                        if s["records"]["data"][idx2+z]["expiryDate"] == last_thursday.strftime("%d-%b-%Y"):
                            print("changing b20")
                            b[i][20] = s["records"]["data"][idx2+z]["strikePrice"]
                            print("NEW B20",b[i][20])
                            try:
                                last_price = s["records"]["data"][idx2+z]["CE"]["bidprice"]
                            except:
                                last_price = 9999
                                logf.write("\nNo CE for : {0} share{1}".format( str(s["records"]["data"][idx2+z]['strikePrice']),b[i][0]))
                                continue
                                #print("NO PE DATA")
                            if s["records"]["data"][idx2+z]["CE"]["askPrice"]==0:
                                ratio1 = 9999
                            else:
                                ratio1 = (s["records"]["data"][idx2+z]["CE"]["askPrice"]- s["records"]["data"][idx2+z]["CE"]["bidprice"])/s["records"]["data"][idx2+z]["CE"]["askPrice"]
                            ratio2 = round(s["records"]["data"][idx2+z]["CE"]["askPrice"]- s["records"]["data"][idx2+z]["CE"]["bidprice"],2)
                            print("BID PRICE , RATIO1, RATIO2")
                            print(last_price,ratio1,ratio2)

                            if ratio1<=0.15 :
                                b[i][21]= last_price
                                cols_color2[i] = count
                                diff22[i][count-1] = ratio2
                                bid2[i][count-1] = s["records"]["data"][idx2+z]["CE"]["bidprice"]
                                ask2[i][count-1] = s["records"]["data"][idx2+z]["CE"]["askPrice"]
                                sp2[i][count-1] = s["records"]["data"][idx2+z]["CE"]["strikePrice"]
                                break
                                            
                            else:
                                diff22[i][count-1] = ratio2
                                bid2[i][count-1] = s["records"]["data"][idx2+z]["CE"]["bidprice"]
                                ask2[i][count-1] = s["records"]["data"][idx2+z]["CE"]["askPrice"]  
                                sp2[i][count-1] = s["records"]["data"][idx2+z]["CE"]["strikePrice"]  
                                print("Cant be done in turn",count)
                                print("OLD B20",b[i][20])
                                #print(diff22)
                                cols_color2[i] = count+1
                                result = np.all(diff22[i] == diff22[i][0])
                                if result:
                                    b[i][20] = s["records"]["data"][int(idx2)]["strikePrice"]
                                    m = np.mean([bid2[i][0],ask2[i][0]])
                                    b[i][21] = float('%.3f'%m)
                                    
                                else:
                                    diff22[i][diff22[i]<= 0 ] = 9999
                                    b[i][20] = sp2[i][np.argmin(diff22[i])]
                                    m1 = np.mean([ask2[i][np.argmin(diff22[i])],bid2[i][np.argmin(diff22[i])]])
                                    b[i][21] = float('%.3f'%m1)
                                    print("avg is",b[i][21])
                                    print("DIFF is ",diff22[i])
                                    print("strike price selected is",sp2[i][np.argmin(diff22[i])])
                                    print("ask and bid are",ask2[i][np.argmin(diff22[i])],bid2[i][np.argmin(diff22[i])])
                                break
    b[i][17] = round(b[i][17],2)
    b[i][21] = round(b[i][21],2)                            
    b[i][18] = round((float(b[i][17]) / float (b[i][16]))*100,2)
    
    b[i][19] = round((b[i][17]/b[i][16]/int(margin)/int(days_left))*3650000,2)
    b[i][22]=round((b[i][21] / float(b[i][20]))*100, 2)
    b[i][23]=round((b[i][21]/b[i][20]/int(margin)/int(days_left))*3650000,2)
    print(b[i][17],b[i][16],int(margin),int(days_left))
    print(b[i][21],b[i][20],int(margin),int(days_left))
    print("BI19,bi21 are",b[i][19],b[i][23])
               
                        
print("turns done",cols_color1)
print("DIff11",diff11)
print(b[i][16],b[i][17])
print("BID1",bid1)
print("ASK1",ask1)
print(last_thursday)
    
print("TURNS TAKEN CE",cols_color2)
print("DIFF22",diff22)
print(b[i][20],b[i][21])
print("BID2",bid2)
print("ASK2",ask2)
print(last_thursday)


directory = os.getcwd()
t='3month'+cd[0]+cd[1]+cd[2]
t
filename = t+'.xlsx'
filename

if filename not in os.listdir(directory):
    df_minmax = pd.DataFrame(columns=['Close_min','Close_max'])
    for i in range(num_shares):
        stk_90 = gh(symbol=stock_names[i], start=date(int(l90d[2]), int(l90d[1]), int(
                l90d[0])), end=date(int(yd[2]), int(yd[1]), int(yd[0])))
        last90=pd.DataFrame(stk_90)
#         print(last90)
        last90min = last90['Close'].min()    
        last90max = last90['Close'].max()    
        b[i][14] = last90min
        b[i][15] = last90max
        df_minmax.loc[i] = [last90min, last90max] 
    df_minmax.to_excel(filename)
    y=yd[0]+yd[1]+yd[2]
    yesterday_file=y+'.xlsx'
    try:
        os.remove(yesterday_file)
    except:
        logf.write("\nFile {0} does not exist to remove".format(str(filename)))
        #print("File does not exist to remove")
        
    
else :
    df_minmax2=pd.read_excel(filename)
    for i in range(num_shares):
        try:
            b[i][14] = df_minmax2['Close_min'][i]
            b[i][15] = df_minmax2['Close_max'][i]
        except:
            logf.write("\nNumber of shares not same in 3 month file saved")
            #print("Number of shares not same in 3 month file saved")

from datetime import datetime


def mdy_to_ymd(d):
    return datetime.strptime(d, '%d,%m,%Y').strftime('%Y-%m-%d')

header = ["Scrip",timestamp,mdy_to_ymd(yesterdayStr), mdy_to_ymd(daybefyesterdayStr), "Nearest Option", "Sell CE @", "Sell PE @", "Premium Total", "Ab IRR", "IRRpa", "Safety Cdr1", "Safety Cdr2", "LowSide", "High Side", "3month Low","3 month High", "F Sell PE of","F SellPE @", "F Ab IRR", "F IRRpa", "F Sell CE of", "F sSell CE  @", " F Ab IRR", "F IRRpa" ]

with open("output.csv", "w+") as my_csv:
    csvWriter=csv.writer(my_csv, delimiter=',')
    
    csvWriter.writerow(header)
    csvWriter.writerows(b)
my_csv.close()


df=pd.read_csv(r'output.csv',index_col=False)
ratio=[]
for i in range(len(df['3month Low'])):
    if df['Sell PE @'][i] != 0:
         rat = round(df['Sell CE @'][i]/df['Sell PE @'][i],2)
    else:
        rat = 9999

    ratio.append(rat)
# df['Ratio']=ratio
df.insert(7, 'Ratio',ratio)
df['Ratio'].fillna(9999)
file = 'output_'+td+'.xlsx'
df.to_excel(file,index=False,sheet_name = 'Output')


wb = openpyxl.load_workbook(file)
sheets=wb.sheetnames
sheets
sh1=wb['Output']
#setting font size
fontStyle = Font(size = "18")
fontStyleH = Font(size = "15", bold= True)
for row in range(2,num_shares+2):
    for col in range(1,len(header)+2):
        sh1.cell(row = row, column = col).font = fontStyle
        sh1.cell(row = 1, column = col).font = fontStyleH

#format cells with word wrap and top alignment    
for row in sh1.iter_rows():  
    for cell in row:      
        cell.alignment = Alignment(wrap_text=True)

#fixing column size
for col in range(1,len(header)+2):
    sh1.column_dimensions[get_column_letter(col)].auto_size = True

#adding borders

border_l = Border(left=Side(border_style=BORDER_THICK, color='00000000'))
border_r =Border(right=Side(border_style=BORDER_THICK, color='00000000'))
border_br =Border(bottom = Side(border_style=BORDER_THICK, color='00000000'),right=Side(border_style=BORDER_THICK, color='00000000'))
border_b = Border(bottom = Side(border_style=BORDER_THICK, color='00000000'))
for row in range(1,num_shares+2):
    sh1.cell(row=row, column=1).border = border_r
    sh1.cell(row=row, column=4).border = border_r
    sh1.cell(row=row, column=11).border = border_r
    sh1.cell(row=row, column=13).border = border_r
    sh1.cell(row=row, column=15).border = border_r
    sh1.cell(row=row, column=17).border = border_r
    sh1.cell(row=row, column=21).border = border_r
    sh1.cell(row=row, column=25).border = border_r

for col in range(1,len(header)+2):
    if col in (1,4,11,13,15,17,21,25):
        sh1.cell(row=num_shares+1, column=col).border = border_br
        sh1.cell(row=1, column=col).border = border_br
    else:
        sh1.cell(row=num_shares+1, column=col).border = border_b
        sh1.cell(row=1, column=col).border = border_b

    

#setting zoom scale
sh1.sheet_view.zoomScale = 53
data=0
color_col=['B','C','H','K','U','Y','N','O','S','R','V','W']
# print(data)
rows = sh1.max_row
# print(rows)
val=rows+1
for j in color_col:
    
#     data = sh1[j].value
    if j== 'H':
        for i in range(2,val):
            var=j+str(i)
            if pd.isna(sh1[var].value)==True:
                pass
            else:

                if float(sh1[var].value) >=1.20 and float(sh1[var].value) <100:
                    fill_var=sh1[var]
                    
                    fill_var.fill = PatternFill(fill_type='solid', start_color='6AA84F', end_color='6AA84F')  # GREEN

                    #wb.save('Names.xlsx')
                elif float(sh1[var].value)<=0.80:
                    fill_var=sh1[var]
                    fill_var.fill = PatternFill(fill_type='solid', start_color='E06666', end_color='E06666')   #RED
                    #wb.save('Names.xlsx')
                else:
                    continue
        
    if j== 'K':
        for i in range(2,val):
            var=j+str(i)
            if sh1[var].value >=200.0:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='6AA84F', end_color='6AA84F')  # GREEN

                #wb.save('Names.xlsx')
            elif sh1[var].value>=150.0 and sh1[var].value<200.0 :
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='FFE599', end_color='FFE599')   # YELLOW
                #wb.save('Names.xlsx')
            else:
                continue
                
    if j=='N':
        for i in range(2,val):
            var=j+str(i)
            if sh1[var].value >= 9:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='6AA84F', end_color='6AA84F')  #green
            elif sh1[var].value < 9 and sh1[var].value >= 6:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='FFE599', end_color='FFE599')  #yellow

            else:
                continue
                
    if j=='O':
        for i in range(2,val):
            var=j+str(i)
            if sh1[var].value >= 9:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='6AA84F', end_color='6AA84F')  #green
            elif sh1[var].value < 9 and sh1[var].value >= 6:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='FFE599', end_color='FFE599')  #yellow
            else:
                continue
                
    if j=='S':
        for i in range(2,val):
            var=j+str(i)
            if cols_color1[i-2]==2:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='FC6247', end_color='FC6247') #Light red
            elif cols_color1[i-2]==3:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='FC6247', end_color='FC6247') #LR
            elif cols_color1[i-2]==4:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='F20C0C', end_color='F20C0C') #DR
            elif cols_color1[i-2]==0:
                share = sh1["A"+str(i)].value
                logf.write("\nError in getting far away PE data of {0}".format(str(share)))
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='D2B48C', end_color='D2B48C') #TAN
            
            else:
                continue
    if j=='R':
        for i in range(2,val):
            var=j+str(i)
            if cols_color1[i-2]==2:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='FC6247', end_color='FC6247') #light red
            elif cols_color1[i-2]==3:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='FC6247', end_color='FC6247') #LR
            elif cols_color1[i-2]==4:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='F20C0C', end_color='F20C0C') #DR
            elif cols_color1[i-2]==0:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='D2B48C', end_color='D2B48C') #TAN
            
            else:
                continue
                
    if j=='V':
        for i in range(2,val):
            var=j+str(i)
            if cols_color2[i-2]==2:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='FC6247', end_color='FC6247') #light red
            elif cols_color2[i-2]==3:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='FC6247', end_color='FC6247') #LR
            elif cols_color2[i-2]==4:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='F20C0C', end_color='F20C0C') #DR
            elif cols_color2[i-2]==0:
                share = sh1["A"+str(i)].value
                logf.write("\nError in getting far away CE data of {0}".format(str(share)))
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='D2B48C', end_color='D2B48C') #TAN
            
            else:
                continue
    if j=='W':
        for i in range(2,val):
            var=j+str(i)
            if cols_color2[i-2]==2:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='FC6247', end_color='FC6247') #light red
            elif cols_color2[i-2]==3:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='FC6247', end_color='FC6247') #LR
            elif cols_color2[i-2]==4:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='F20C0C', end_color='F20C0C') #DR
            elif cols_color2[i-2]==0:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='D2B48C', end_color='D2B48C') #TAN
            
            else:
                continue
    
                
                
                
    if j=='U':
        for i in range(2,val):
            var=j+str(i)
            if sh1[var].value >= 48.0:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='3E8D1C', end_color='3E8D1C')    # DARK GREEN
                #wb.save('Names.xlsx')
            elif sh1[var].value>=36.0 :
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='77F740', end_color='77F740')     # GREEN
                #wb.save('Names.xlsx')       
            elif sh1[var].value >=24.0:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='FFE599', end_color='FFE599')      # YELLOW
                #wb.save('Names.xlsx')
            else:
                continue
        

    if j=='Y':
        for i in range(2,val):
            var=j+str(i)
            if sh1[var].value >= 48.0:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='3E8D1C', end_color='3E8D1C')    # DARK GREEN
                #wb.save('Names.xlsx')
            elif sh1[var].value>=36.0 :
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='77F740', end_color='77F740')     # GREEN
                #wb.save('Names.xlsx')       
            elif sh1[var].value >=24.0:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='FFE599', end_color='FFE599')      # YELLOW
                #wb.save('Names.xlsx')
            else:
                continue            
    
    if j=='B':
        for i in range(2,val):
            var=j+str(i)
            try:
                perc = ((sh1[var].value - sh1['C'+str(i)].value)/sh1['C'+str(i)].value)*100
            except:
                perc = 0
            if perc >=1 and perc <2:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='B6D7A8', end_color='B6D7A8')   # LIGHT GREEN
                #wb.save('Names.xlsx')
            elif perc >=2 and perc <4:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='6AA84F', end_color='6AA84F')     # GREEN
                #wb.save('Names.xlsx')       
            elif perc >4:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='3E8D1C', end_color='3E8D1C')     # DARK GREEN
                #wb.save('Names.xlsx')
            elif perc >-2 and perc <=-1:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='F4CCCC', end_color='F4CCCC')     # LIGHT RED
                #wb.save('Names.xlsx')
            elif perc >-4 and perc <=-2:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='E06666', end_color='E06666')     # RED
                #wb.save('Names.xlsx')
            elif perc <-4:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='D93737', end_color='D93737')      # DARK RED
                #wb.save('Names.xlsx')
            else:
                continue 

    if j=='C':
        for i in range(2,val):
            var=j+str(i)
            try:
                perc = ((sh1[var].value - sh1['D'+str(i)].value)/sh1['D'+str(i)].value)*100
            except:
                perc = 0
            if perc >=1 and perc <2:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='B6D7A8', end_color='B6D7A8')
                #wb.save('Names.xlsx')
            elif perc >=2 and perc <4:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='6AA84F', end_color='6AA84F')
                #wb.save('Names.xlsx')       
            elif perc >4:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='3E8D1C', end_color='3E8D1C')
                #wb.save('Names.xlsx')
            elif perc >-2 and perc <=-1:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='F4CCCC', end_color='F4CCCC')
                #wb.save('Names.xlsx')
            elif perc >-4 and perc <=-2:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='E06666', end_color='E06666')
                #wb.save('Names.xlsx')
            elif perc <-4:
                fill_var=sh1[var]
                fill_var.fill = PatternFill(fill_type='solid', start_color='D93737', end_color='D93737')
                #wb.save('Names.xlsx')
            else:
                continue  
        
one_d = ['H','I','J','K','L','M','N','O','P','Q','T','U','X','Y']

for i in one_d:
    for j in range(2,val):
        try:
            sh1[i+str(j)].value = float(sh1[i+str(j)].value)
            sh1[i+str(j)] = round(sh1[i+str(j)].value,1)
        except:
            pass

colname=['J','K', 'N','O','T','U','X','Y']
for i in colname:
    for j in range(2,val):
        sh1[i+str(j)] = str(sh1[i+str(j)].value)+'%'
        
wb.save(file)
